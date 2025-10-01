"""
程序功能：帮助我们自动整理电脑生成的PDF发票文件。PS:暂时不支持扫描生成的PDF文件

author：zpeng
email：zpeng0614@gmail.com

"""
from openpyxl import Workbook
from PyPDF2 import PdfMerger
from datetime import datetime
import pymupdf as pdf
import pytesseract
import pdfplumber
import os
import re

#这个是对advanced的分支的修改
def rename_pdf():
    """对所有的单个pdf文件按照金额进行重命名"""
    #获取单张发票中的数据
    for i in range (invoice_num):
        text = ''#定义一个接收发票数据的控字符串
        tem_list1 = []
        tem_list2 = []
        with pdf.open(os.listdir()[i]) as file:
            for j in file:
                tem_text = j.get_text()
                text += tem_text
                tem_text.strip()

        #筛选出单张发票中金额最大的且带有人民币符号的即为发票金额
        if '¥' in text:
            tem_list1 = text.split('\n')
            for k in range(len(tem_list1)):
                if '¥' in tem_list1[k]:
                    num_str = tem_list1[k].strip("(小写) ").strip().strip("¥").strip()
                    tem_list2.append(float(num_str))
        elif '￥' in text:
            tem_list1 = text.split('\n')
            for k in range(len(tem_list1)):
                if '￥' in tem_list1[k]:
                    num_str = tem_list1[k].strip("(小写) ").strip().strip('￥').strip()
                    tem_list2.append(float(num_str))
        #图片扫描获得的pdf文件暂时不支持
        elif text == '':
            print("第",(i+1),"张名称为",os.listdir()[i],"的发票是扫描版pdf,而非电脑生成pdf,此脚本暂时不支持这种模式",sep = '')#表示这不是电脑生成的pdf，而是扫描版的pdf无法提取出内容
            print('再次声明该脚本暂时不支持该发票文件，该发票文件也不会计入总pdf文件及总金额之中，请在https://github.com/ZPENG0614/invoice_auto.git中提出issues或是自行修改后提出PR',sep = '')
            tem_list2 = [888888888]
        else:
            print("第",(i+1),"张名称为",os.listdir()[i],"发生未知异常,请在https://github.com/ZPENG0614/invoice_auto.git中提出issues或是自行修改后提出PR")
            tem_list2 = [888888888]

        #对于所有支持的文件（电脑生成的PDF）进行重命名
        if text != '':
            current_time = datetime.now()
            current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S.%f")
            os.rename(os.listdir()[i],str(max(tem_list2))+'    '+str(current_time)+'.pdf')
            pdfs_num.append(str(max(tem_list2))+'    '+str(current_time))
            pdfs_money.append(max(tem_list2))

def merge_pdf():
    """合并重命名之后的单个pdf文件"""

    # 创建合并器对象
    merger = PdfMerger()
    pdfs_list = os.listdir()
    # 用于存储（数字, 文件名）的列表
    sorted_items = []

    #正则表达式
    for line in pdfs_list:
        match = re.match(r'^(\d+\.?\d*)\s+(.*)$', line)
        if match:
            number = float(match.group(1))
            filename = match.group(2)
            sorted_items.append((number, filename))

    # 按数字升序排序
    sorted_items.sort(key=lambda x: x[0])

    # 提取排序后的文件名
    sorted_filenames = [str(item[0])+'    '+item[1] for item in sorted_items]

    for filename in sorted_filenames:
        merger.append(filename)

    os.chdir('..')# 切换到上一级目录
    # 保存合并后的文件
    merger.write("合并后pdf文件.pdf")
    merger.close()


def excel_pdf():
    """创建Excel文件并写入数据"""
    # 创建工作簿对象
    wb = Workbook()
    # 获取默认的活动工作表（第一个工作表）
    ws = wb.active

    # 给工作表命名
    ws.title = "发票金额汇总"
    ws['A1'] = '发票金额'
    ws['C3'] = '发票张数'
    ws['C4'] = '发票总金额'

    #依次写入发票的金额数据
    for i in range(len(pdfs_num)):
        ws['A'+str(i+2)] = str(pdfs_money[i])

    ws['D3'] = str(invoice_num+1)#发票张数
    ws['D4'] = str(sum(pdfs_money))#发票总金额
    # 保存文件
    wb.save('发票金额数据.xlsx')


if __name__ == '__main__':
#上一行代码决定了这段程序的运行方式。
#1、直接运行，内置变量__name__将被赋值为__main__,条件成立直接执行
#2、作为模块导入到其他的模块之中，内置变量__name__将被赋值为文件名（不加.py），条件不成立

    #程序开始提示信息
    print('\n')
    print("使用方法：将所有发票pdf文件放到名为发票的文件夹中，然后将发票助手与发票文件夹放入同一级目录下，运行程序即可")
    print("说明：该程序仅支持电脑自动生成的PDF，不支持扫描的PDF，对于扫描的PDF该程序不会进行任何操作。PS:持续更新中......")
    print("按enter键开始运行：entering.........")
    print("按q键退出.........")

    user_input = input()#获取用户输入
    if user_input == '':
        print("程序运行中.......")
        print('\n')

        #一些全局变量
        current_path = os.getcwd()  # 获取当前脚本所在目录
        os.chdir("发票")  # 进入发票目录，以提取相关信息
        invoice_num = len(os.listdir())  # 获取当前发票文件的数量
        pdfs_num = []  # 用来存储发票名称的列表
        pdfs_money = []  # 用来存储发票金额的列表
        try:
            #对PDF文件进行操作的函数
            rename_pdf()  # 对所有的发票重命名

        except PermissionError:
            print('发票pdf文件已经在其他程序中打开，请关闭后重新运行程序')
        else:
            merge_pdf()  # 合并所有的发票
            excel_pdf()  # 将发票数据写入exl文件
        finally:
            #程序结束提示信息
            print('\n')
            input("按回车键退出>>>>>>")
    elif user_input == 'q':
        print("程序关闭")

